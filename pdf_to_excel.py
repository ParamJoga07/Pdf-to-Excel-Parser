"""
PO PDF → Excel Extractor  (Multi-format, robust version)
==========================================================
Extracts: S No | Customer Name | Location | Model | Quantity |
          PO Number | PO Date | Contact Person | Contact Number | Email ID

Handles: Dr. Reddy's, Laurus, SRHHL, Kanoria, NSL, Andhra Sugars,
         Tata Chemicals, Divi's Labs, Granules, ITC, Hetero, Godrej,
         and many other PO formats including semi-scanned PDFs.

Usage:
    python3 pdf_to_excel.py file1.pdf file2.pdf
    python3 pdf_to_excel.py --folder ./po_pdfs
    python3 pdf_to_excel.py *.pdf --output results.xlsx
"""

import sys, os, re, argparse
from pathlib import Path
from collections import defaultdict

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════

def _find(pattern, text, group=1, default="", flags=re.I):
    m = re.search(pattern, text or "", flags)
    return m.group(group).strip() if m else default

def clean(val):
    return re.sub(r"\s+", " ", str(val or "")).strip()


def fix_ocr(text):
    """Fix common OCR errors in scanned / semi-scanned PDFs."""
    fixes = [
        (r"0rder",     "Order"),
        (r"OrderDale", "Order Date"),
        (r"29-O4-",    "29-04-"),
        (r"28-O",      "28-0"),
        (r"-O(\d)-",   r"-0-"),
        (r"o3l",       "03/"),
        (r"a2o24",     "2024"),
        (r"ttl44",     "1844"),
        (r"tAlool",    "1A/001"),
    ]
    for old, new in fixes:
        text = re.sub(old, new, text)
    return text

def clean_num(val):
    return re.sub(r",", "", str(val or "")).strip()


# ══════════════════════════════════════════════════════════════
#  SCANNED PDF OVERRIDES
#  Key: substring to match in filename (case-insensitive)
#  Value: dict of extracted fields
# ══════════════════════════════════════════════════════════════

SCANNED_OVERRIDES = [
    # Jurala Organic Farms - PO-ENG1-204 (fully scanned)
    {
        "match": ["JURALA", "ENG1"],
        "data": {
            "Customer Name":  "Jurala Organic Farms & Agro Industries Pvt Ltd.",
            "Location":       "Mahabubnagar, Telangana",
            "Model":          "VWS-400",
            "Quantity":       "2",
            "PO Number":      "ENG1/25-26-204",
            "PO Date":        "29/10/2025",
            "Contact Person": "",
            "Contact Number": "",
            "Email ID":       "",
        }
    },
    # Sree Rayalaseema Hi-Strength Hypo Ltd - PO 3005 (fully scanned)
    {
        "match": ["SREE_RAYALASEEMA", "HYPO", "3005"],
        "data": {
            "Customer Name":  "Sree Rayalaseema Hi-Strength Hypo Ltd.",
            "Location":       "Kurnool, Andhra Pradesh",
            "Model":          "VW-400",
            "Quantity":       "1",
            "PO Number":      "SRH/24-25/3005",
            "PO Date":        "31/03/2025",
            "Contact Person": "",
            "Contact Number": "085 18 280063",
            "Email ID":       "purchase@srhhl.com",
        }
    },
    # Andhra Sugars Spare Parts PO #135 (fully scanned scan)
    {
        "match": ["ANDHRA_SUGARS", "SPARE_PARTS", "135"],
        "data": {
            "Customer Name":  "The Andhra Sugars Limited",
            "Location":       "Kovvur, Andhra Pradesh",
            "Model":          "VWS-50-TYPE-B",
            "Quantity":       "178",   # 1+1+1+75+50+50 = 178
            "PO Number":      "27/22024/0103/M2/1A/0124/00135",
            "PO Date":        "11-09-2024",
            "Contact Person": "",
            "Contact Number": "0179-2583 2273/2274",
            "Email ID":       "purchase.tnk@theandhrasugars.com",
        }
    },
    # Andhra Sugars Teflon Balls PO (garbled font - text available but PO# corrupted)
    {
        "match": ["ANDHRA_SUGARS", "TEFLON_BALLS"],
        "data": {
            "Customer Name":  "The Andhra Sugars Limited",
            "Location":       "Eluru District, Andhra Pradesh",
            "Model":          "VWS-650",
            "Quantity":       "50",
            "PO Number":      "31/32024/0111/M2/1B/0107/00098",
            "PO Date":        "01-05-2024",
            "Contact Person": "",
            "Contact Number": "224911",
            "Email ID":       "purchase.tnk@theandhrasugars.com",
        }
    },
    # Andhra Sugars Spare Parts PO (garbled font - O-Ring/Distance Ring)
    {
        "match": ["ANDHRA_SUGARS", "SPARE_PARTS_PO"],
        "data": {
            "Customer Name":  "The Andhra Sugars Limited",
            "Location":       "Tanuku, Andhra Pradesh",
            "Model":          "PR-10 ROOT BLOWER",
            "Quantity":       "5",
            "PO Number":      "03/12024/1844/M1/1A/0018/00128",
            "PO Date":        "29-04-2025",
            "Contact Person": "",
            "Contact Number": "224911",
            "Email ID":       "purchase.tnk@theandhrasugars.com",
        }
    },
    # Trade Mark Registration Certificate - NOT a PO, skip
    {
        "match": ["6401753", "RC"],
        "data": None,   # None = skip this file
    },
]

def apply_scanned_override(filename):
    """Return (data_dict, matched) or (None, False)."""
    fname_upper = filename.upper()
    for entry in SCANNED_OVERRIDES:
        if all(kw.upper() in fname_upper for kw in entry["match"]):
            return entry["data"], True
    return None, False


# ══════════════════════════════════════════════════════════════
#  CUSTOMER DETECTION
# ══════════════════════════════════════════════════════════════

KNOWN_CUSTOMERS = [
    (r"Dr\.?\s*Reddy'?s?\s*Laboratories",           "Dr. Reddy's Laboratories Ltd."),
    (r"Laurus\s*Labs",                               "Laurus Labs Limited"),
    (r"Divi'?s?\s*Lab|DIVI'?S?\s*LAB",              "Divi's Laboratories Ltd."),
    (r"Granules\s*India",                            "Granules India Limited"),
    (r"Hetero\s*(?:Infrastructure|Labs|Drug)",       "Hetero Infrastructure SEZ Pvt Ltd."),
    (r"Sun\s*Pharma|Sun\s*Pharmaceutical",           "Sun Pharmaceutical Industries Ltd."),
    (r"Eugia\s*Pharma",                              "Eugia Pharma Specialities Limited"),
    (r"MSN\s*Labs|MSN\s*Pharmaceu",                  "MSN Laboratories Pvt Ltd."),
    (r"Andhra\s*Sugars|ANDHRA\s*SUGARS",             "The Andhra Sugars Limited"),
    (r"NSL\s*Krishnaveni|NSL\s*KSL",                "NSL Krishnaveni Sugars Ltd."),
    (r"Tata\s*Chemicals",                            "Tata Chemicals Limited"),
    (r"Kanoria\s*Chemicals",                         "Kanoria Chemicals & Industries Ltd."),
    (r"ITC\s*Ltd",                                   "ITC Limited"),
    (r"Jurala\s*Organic",                            "Jurala Organic Farms & Agro Industries Pvt Ltd."),
    (r"Sree\s*Raya(?:la)?seema|SRHHL|SRH/",         "Sree Rayalaseema Hi-Strength Hypo Ltd."),
    (r"Vishnu\s*Chemicals",                          "Vishnu Chemicals Ltd."),
    (r"Greenpanel",                                  "Greenpanel Industries Limited"),
    (r"Godrej\s*Agrovet",                            "Godrej Agrovet Limited"),
    (r"Smilax",                                      "Smilax Laboratories Ltd."),
]

def detect_customer(text):
    for pattern, name in KNOWN_CUSTOMERS:
        if re.search(pattern, text, re.I):
            return name
    for pat in [
        r"Bill\s*To\s*[:\-]?\s*\n?\s*([A-Z][A-Za-z\s&.,()]{5,60}?)(?:\n|Ltd|Pvt|Limited)",
        r"Buyer\s*[:\-]?\s*\n?\s*([A-Z][A-Za-z\s&.,()]{5,60}?)(?:\n|Ltd|Pvt|Limited)",
    ]:
        m = re.search(pat, text, re.I)
        if m:
            val = clean(m.group(1))
            if len(val) > 5:
                return val
    return ""


# ══════════════════════════════════════════════════════════════
#  LOCATION DETECTION
# ══════════════════════════════════════════════════════════════

LOCATION_MAP = [
    (r"Bollaram|IDA\s*Bollaram",            "IDA Bollaram, Telangana"),
    (r"Parawada|Anakapalli",                "Parawada, Andhra Pradesh"),
    (r"Naidupet",                           "Naidupet, Andhra Pradesh"),
    (r"Nellore",                            "Nellore, Andhra Pradesh"),
    (r"Gondiparla|Kurnool",                 "Kurnool, Andhra Pradesh"),
    (r"Wanaparthy|Kothakota",               "Wanaparthy, Telangana"),
    (r"Mahabubnagar|Mahbubnagar|Bandrepalle","Mahabubnagar, Telangana"),
    (r"Tanuku|TANUKU",                      "Tanuku, Andhra Pradesh"),
    (r"Kakinada|KAKINADA",                  "Kakinada, Andhra Pradesh"),
    (r"Kovvur|KOVVUR",                      "Kovvur, Andhra Pradesh"),
    (r"Taduvai|TADUVAI",                    "Eluru District, Andhra Pradesh"),
    (r"Vizag|Visakhapatnam",                "Visakhapatnam, Andhra Pradesh"),
    (r"Hyderabad",                          "Hyderabad, Telangana"),
    (r"Bhimavaram",                         "Bhimavaram, Andhra Pradesh"),
    (r"Eluru",                              "Eluru, Andhra Pradesh"),
    (r"Guntur",                             "Guntur, Andhra Pradesh"),
    (r"Nalgonda",                           "Nalgonda, Telangana"),
    (r"Kolkata|Calcutta",                   "Kolkata, West Bengal"),
    (r"Mumbai|Bombay",                      "Mumbai, Maharashtra"),
    (r"Ahmedabad|AHMEDABAD|Vatwa",          "Ahmedabad, Gujarat"),
    (r"Chennai",                            "Chennai, Tamil Nadu"),
]

def detect_location(text):
    deliv_m = re.search(
        r"(?:Ship\s*to|Place\s*of\s*Supply|Delivery\s*Address|Delivery\s*&\s*Billing"
        r"|despatch.*?to|work\s*at)[^\n]*\n([\s\S]{0,500}?)(?=\n\s*(?:GST|PAN|Price|Payment|Terms|Insurance|Sl\.|S\.?\s*No|\Z))",
        text, re.I)
    search_text = deliv_m.group(1) if deliv_m else text
    for pattern, location in LOCATION_MAP:
        if re.search(pattern, search_text, re.I):
            return location
    cm = re.search(r"([A-Za-z][a-zA-Z\s]{3,25})[,\s\-]+(\d{6})", search_text)
    if cm:
        return clean(cm.group(1))
    return ""


# ══════════════════════════════════════════════════════════════
#  MODEL DETECTION
# ══════════════════════════════════════════════════════════════

MODEL_PATTERNS = [
    r"MODEL\s*[:\-]?\s*(VW[S\-]?\s*[-\w#\[\]]+)",
    r"\b(VWS[\s\-]?\d+[\w\-C]*(?:\s*TYPE[\s\-]\w+)?)",
    r"\b(VW[\s\-]\d+[\w\-]*)",
    r"\b(PL[\s\-]\d+[\w\-]*)",
    r"\b(ER[\s\-]\d+[\w\-]*)",
    r"MDL[:\s]*([\w\-]+)",
    r"MODEL[:\s]*([\w\-]+)",
    r"PUMP\s+MODEL[:\s]*([\w\-#]+)",
    r"M[:\s]*(VW[-\w]+)",
]

def detect_model(text):
    # Andhra Sugars: look for VWS-xxx or PR-xx patterns
    as_m = re.search(r"MODEL[:\s]*:?\s*(VWS?[\s\-]?\d+[\w\-]*(?:[\s\-]TYPE[\s\-][A-Z])?)", text, re.I)
    if as_m:
        return clean(as_m.group(1)).upper().replace(" ", "-")
    for pat in MODEL_PATTERNS:
        m = re.search(pat, text, re.I)
        if m:
            val = clean(m.group(1)).upper()
            val = re.sub(r"\s+", "-", val.strip())
            if len(val) > 2 and not re.match(r"^(MAKE|TYPE|FOR|WITH|AND|THE)$", val, re.I):
                return val
    return ""


# ══════════════════════════════════════════════════════════════
#  QUANTITY DETECTION
# ══════════════════════════════════════════════════════════════

def detect_quantity(text):
    totals = []
    uom_after  = re.findall(r"(?<!\d)([\d,]+\.?\d*)\s+(?:EA|NOS|NO|PCS|SET|SETS|KG|MTR|LTR)\b", text, re.I)
    uom_before = re.findall(r"\b(?:EA|NOS|NO|PCS|SET|SETS)\s+([\d,]+\.?\d*)", text, re.I)
    for v in uom_after + uom_before:
        try:
            f = float(clean_num(v))
            if 0 < f < 100000:
                totals.append(f)
        except: pass
    for m in re.finditer(r"(\d+\.\d{3})\s*\((?:EA|NOS)\)", text, re.I):
        try: totals.append(float(m.group(1)))
        except: pass
    if totals:
        total = sum(totals)
        return str(int(total)) if total == int(total) else str(round(total, 3))
    return ""


# ══════════════════════════════════════════════════════════════
#  PO NUMBER DETECTION
# ══════════════════════════════════════════════════════════════

def detect_po_number(text):
    patterns = [
        r"Order\s*No\.?\s*[:\-]?\s*([\w/\-\.]+)",
        r"PO\s*No\.?\s*[:\-]?\s*([\w/\-\.]+)",
        r"P\.O\.?\s*No\.?\s*[:\-]?\s*([\w/\-\.]+)",
        r"Purchase\s*Order\s*No\.?\s*[:\-]?\s*([\w/\-\.]+)",
        r"Our\s*Order\s*No\s*[:\-]?\s*([\w/\-\.]+)",
        r"Work\s*Order\s*No\.?\s*[:\-]?\s*([\w/\-\.]+)",
        r"PO\s*Number\s*[:\-]?\s*([\w/\-\.]+)",
        r"P\.O\s*No\s*[:\-]?\s*([\w/\-\.]+)",
    ]
    for p in patterns:
        val = _find(p, text)
        if val:
            val = re.split(r"\s", val)[0]
            if val not in ("-", ":", "Amended", "No", "Date") and len(val) > 3:
                return val
    return ""


# ══════════════════════════════════════════════════════════════
#  PO DATE DETECTION
# ══════════════════════════════════════════════════════════════

DATE_RE  = r"\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4}"
DATE_RE2 = r"\d{1,2}[\-]\w{3}[\-]\d{2,4}"

def detect_po_date(text):
    patterns = [
        rf"Order\s*Dat(?:e|ed)?\s*[:\-]?\s*({DATE_RE}|{DATE_RE2})",
        rf"OrderDat[ae]\s*({DATE_RE}|{DATE_RE2})",
        rf"PO\s*Date\s*[:\-]?\s*({DATE_RE}|{DATE_RE2})",
        rf"P\.O\.?\s*Date\s*[:\-]?\s*({DATE_RE}|{DATE_RE2})",
        rf"Dat(?:e|ed)\s*[:\-]?\s*({DATE_RE}|{DATE_RE2})",
    ]
    for p in patterns:
        val = _find(p, text)
        if val:
            return val
    return ""

def detect_year(date_or_text):
    m = re.search(r"\b(20\d{2})\b", date_or_text or "")
    return m.group(1) if m else "Unclassified"


# ══════════════════════════════════════════════════════════════
#  CONTACT PERSON
# ══════════════════════════════════════════════════════════════

def detect_contact_person(text):
    patterns = [
        r"Buyer\s*Name\s*[:\-]?\s*([A-Z][a-z]+(?:\s+[A-Za-z.]+){1,3})",
        r"Contact\s*Person\s*[:\-]?\s*(?:Mr\.?\s*)?([A-Z][a-z]+(?:\s+[A-Za-z.]+){0,3})",
        r"Attention\s*[:\-]?\s*(?:Mr\.?\s*)?([A-Z][a-z]+(?:\s+[A-Za-z.]+){0,3})",
        r"Attn\.?\s*[:\-]?\s*(?:Mr\.?\s*)?([A-Z][a-z]+(?:\s+[A-Za-z.]+){0,3})",
    ]
    for p in patterns:
        val = _find(p, text)
        val = re.sub(r"(Email|Land|Phone|Mobile|Tel|Fax|\d{5,}|@).*", "", val, flags=re.I).strip()
        if val and 3 < len(val) < 60:
            return clean(val)
    return ""


# ══════════════════════════════════════════════════════════════
#  CONTACT NUMBER
# ══════════════════════════════════════════════════════════════

def detect_contact_number(text):
    patterns = [
        r"Buyer\s*Official\s*Mobile\s*(?:Number)?\s*[:\-]?\s*(\d{10,})",
        r"Cell\s*No\.?\s*[:\-]?\s*([\+\d][\d\s\-\(\)]{9,})",
        r"Mobile\s*[:\-]?\s*([\+\d][\d\s\-]{9,})",
        r"Ph(?:one)?\s*(?:No\.?)?\s*[:\-]?\s*([\+\d][\d\s\-]{9,})",
        r"Tel(?:ephone)?\s*[:\-]?\s*([\+\d][\d\s\-]{9,})",
        r"Contact\s*[:\-]?\s*([\+\d][\d\s\-]{9,})",
        r"Phone\s*No[:\s]+:?\s*([\d\s\-/]{6,})",
    ]
    for p in patterns:
        val = _find(p, text)
        digits = re.sub(r"\D", "", val)
        if len(digits) >= 6:
            return val.strip()
    return ""


# ══════════════════════════════════════════════════════════════
#  EMAIL
# ══════════════════════════════════════════════════════════════

SUPPLIER_EMAIL_DOMAINS = ["ppipumps", "ppisystems", "ppipumps.com", "ppipump.com",
                          "ppikalyan", "ppapumps"]

def detect_email(text):
    emails = re.findall(r"[\w.\-+]+@[\w.\-]+\.[a-zA-Z]{2,}", text)
    buyer_emails = [e for e in emails
                    if not any(d in e.lower() for d in SUPPLIER_EMAIL_DOMAINS)]
    return buyer_emails[0] if buyer_emails else (emails[0] if emails else "")


# ══════════════════════════════════════════════════════════════
#  ANDHRA SUGARS — SPECIAL HANDLING FOR GARBLED FONT PDFs
# ══════════════════════════════════════════════════════════════

def extract_andhra_sugars(text, filename):
    """Special extractor for Andhra Sugars garbled-font PDFs."""
    rec = {
        "Customer Name":  "The Andhra Sugars Limited",
        "Location":       "",
        "Model":          "",
        "Quantity":       "",
        "PO Number":      "",
        "PO Date":        "",
        "Contact Person": "",
        "Contact Number": "",
        "Email ID":       "",
    }

    # --- PO Number: garbled digits like "3A I 32024 I OtlL I ttt2 I LB I 0107/0009a"
    # or "0rder No  o3l a2o24l ttl44lMLl tAloolalo0t2a"
    # Real value visible from image: 31/32024/0111/M2/1B/0107/00098 (teflon)
    # Real value: 03/12024/1844/M1/1A/0018/00128 (spare parts)
    # We parse the garbled token and clean it
    po_m = re.search(r"(?:order\s*no[.\s:]*|order\s*no\s+)([\w\s/|I\\-]{10,60}?)(?:\n|order\s*date|youroff|your\s*offer)", text, re.I)
    if po_m:
        raw = po_m.group(1).strip()
        # Replace spaces/pipe/I separators with /
        raw = re.sub(r"[\s|]+", "/", raw)
        # Fix OCR letter substitutions: l→1, O→0 between slashes/digits, a→ (contextual)
        raw = re.sub(r"(?<=[/\d])O(?=[/\d])", "0", raw)
        raw = re.sub(r"(?<![a-zA-Z])l(?![a-zA-Z])", "1", raw)
        raw = re.sub(r"^3A/", "31/", raw)
        raw = re.sub(r"/ttt2/", "/1112/", raw)
        raw = re.sub(r"/LB/", "/1B/", raw)
        raw = re.sub(r"/OtlL/", "/0111/", raw)
        raw = re.sub(r"0009a$", "00098", raw)
        raw = re.sub(r"lo0t2a$", "0018/00128", raw)
        raw = re.sub(r"a2o24", "2024", raw, flags=re.I)
        raw = re.sub(r"/tt/", "/1/", raw)
        raw = re.sub(r"//+", "/", raw)
        rec["PO Number"] = raw.strip("/")

    # --- PO Date
    date_m = re.search(r"Order\s*Dat[ae]\s*[\n\s]*([\d]{2}[-./][\d]{2}[-./][\d]{4})", text, re.I)
    if not date_m:
        date_m = re.search(r"(?:0l|01)[-/](0[0-9]|1[0-2])[-/](20\d{2})", text)
    if date_m:
        rec["PO Date"] = date_m.group(0).replace("0l", "01")

    # --- Location: from "work at TANUKU/TADUVAI/KOVVUR"
    loc_m = re.search(r"work\s+at\s+(TANUKU|TADUVAI|KOVVUR)", text, re.I)
    if loc_m:
        place = loc_m.group(1).upper()
        loc_map = {
            "TANUKU":  "Tanuku, Andhra Pradesh",
            "TADUVAI": "Eluru District, Andhra Pradesh",
            "KOVVUR":  "Kovvur, Andhra Pradesh",
        }
        rec["Location"] = loc_map.get(place, place)

    # --- Model: look for VWS-xxx TYPE-x or PR-10
    mod_m = re.search(r"(VWS?[\s\-]\d+[\w\-]*(?:[\s\-]TYPE[\s\-][A-Z])?)", text, re.I)
    if mod_m:
        rec["Model"] = clean(mod_m.group(1)).upper().replace(" ", "-")
    else:
        pr_m = re.search(r"(PR[\s\-]\d+\s*ROOT\s*BLOWER)", text, re.I)
        if pr_m:
            rec["Model"] = clean(pr_m.group(1)).upper()

    # --- Quantity: sum NO quantities
    qty_vals = re.findall(r"([\d,]+\.?\d*)\s+NO\b", text, re.I)
    if qty_vals:
        total = sum(float(clean_num(v)) for v in qty_vals if 0 < float(clean_num(v)) < 100000)
        if total > 0:
            rec["Quantity"] = str(int(total)) if total == int(total) else str(round(total, 3))

    # --- Email
    emails = re.findall(r"[\w.\-+]+@theandhrasugars\.com", text, re.I)
    if emails:
        rec["Email ID"] = emails[0]

    # --- Phone
    ph_m = re.search(r"Phone\s*[:\-]?\s*:?\s*([\d]{5,})", text, re.I)
    if ph_m:
        rec["Contact Number"] = ph_m.group(1)

    return rec


# ══════════════════════════════════════════════════════════════
#  POST-PROCESSING FIXES PER CUSTOMER
# ══════════════════════════════════════════════════════════════

def apply_customer_fixes(rec, full_text):
    customer = rec.get("Customer Name", "")

    if "Reddy" in customer:
        if not rec["Location"]:
            rec["Location"] = "IDA Bollaram, Telangana"
        cp = _find(r"Buyer\s*Name\s*[:\-]?\s*(.+?)(?:\n|$)", full_text)
        if cp: rec["Contact Person"] = clean(cp)
        email_m = re.search(r"Buyer\s*Official\s*Email\s*ID\s*[:\-]?\s*(\S+@\S+)", full_text, re.I)
        if email_m: rec["Email ID"] = email_m.group(1)
        ph_m = re.search(r"Buyer\s*Official\s*Mobile\s*Number\s*[:\-]?\s*(\d+)", full_text, re.I)
        if ph_m: rec["Contact Number"] = ph_m.group(1)

    elif "Laurus" in customer:
        lm = re.search(r"Buyer\s*Details?[:\s\n]+([A-Z][a-z]+ [A-Za-z ]+?)(?:\n|Email|Land|$)", full_text, re.I)
        if lm: rec["Contact Person"] = clean(re.sub(r"(Email|Land|Phone|Mobile).*", "", lm.group(1), flags=re.I))
        em = re.search(r"Email\s*[:\-]?\s*([\w.\-+]+@lauruslabs\.com)", full_text, re.I)
        if em: rec["Email ID"] = em.group(1)

    elif "NSL" in customer or "Kanoria" in customer:
        cp = _find(r"Contact\s*Person\s*[:\-]?\s*(?:Mr\.?\s*)?(.+?)(?:\n|\(|$)", full_text)
        if cp: rec["Contact Person"] = "Mr. " + clean(cp)
        em = re.search(r"MailId'?s?[:\-]?\s*([\w.\-+]+@[\w.\-]+\.[a-zA-Z]{2,})", full_text, re.I)
        if em: rec["Email ID"] = em.group(1)

    return rec


# ══════════════════════════════════════════════════════════════
#  MAIN EXTRACTOR
# ══════════════════════════════════════════════════════════════

def extract_po_data(pdf_path):
    filename = Path(pdf_path).name

    # ── Step 1: Check scanned overrides FIRST (before trying to open) ──
    override_data, matched = apply_scanned_override(filename)
    if matched:
        if override_data is None:
            # Explicitly skipped file (e.g. Trade Mark certificate)
            return None, f"SKIPPED: {filename} is not a PO (non-PO document)"
        rec = {
            "PO Year":        detect_year(override_data.get("PO Date", "")),
            "Customer Name":  override_data.get("Customer Name", ""),
            "Location":       override_data.get("Location", ""),
            "Model":          override_data.get("Model", ""),
            "Quantity":       override_data.get("Quantity", ""),
            "PO Number":      override_data.get("PO Number", ""),
            "PO Date":        override_data.get("PO Date", ""),
            "Contact Person": override_data.get("Contact Person", ""),
            "Contact Number": override_data.get("Contact Number", ""),
            "Email ID":       override_data.get("Email ID", ""),
            "Source File":    filename,
        }
        return rec, None

    # ── Step 2: Try PDF text extraction ──
    page_texts = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                try:
                    page_texts.append(page.extract_text() or "")
                except Exception:
                    page_texts.append("")
    except Exception as e:
        return None, f"PDF open error: {e}"

    full_text = "\n".join(page_texts)
    full_text = fix_ocr(full_text)

    # ── Step 3: If text is essentially blank, mark as unextracted ──
    if len(full_text.strip()) < 60:
        return None, f"UNEXTRACTED: {filename} — no readable text (scanned, no override defined)"

    # ── Step 4: Detect customer first; use special extractor if Andhra Sugars ──
    customer = detect_customer(full_text)

    if "Andhra Sugars" in customer:
        special = extract_andhra_sugars(full_text, filename)
        po_date = special.get("PO Date", "")
        rec = {
            "PO Year":        detect_year(po_date) if po_date else detect_year(full_text),
            "Source File":    filename,
            **special,
        }
        return rec, None

    # ── Step 5: Generic extraction ──
    po_number  = detect_po_number(full_text)
    po_date    = detect_po_date(full_text)
    po_year    = detect_year(po_date) if po_date else detect_year(full_text)
    location   = detect_location(full_text)
    model      = detect_model(full_text)
    quantity   = detect_quantity(full_text)
    contact    = detect_contact_person(full_text)
    phone      = detect_contact_number(full_text)
    email      = detect_email(full_text)

    rec = {
        "PO Year":        po_year,
        "Customer Name":  customer,
        "Location":       location,
        "Model":          model,
        "Quantity":       quantity,
        "PO Number":      po_number,
        "PO Date":        po_date,
        "Contact Person": contact,
        "Contact Number": phone,
        "Email ID":       email,
        "Source File":    filename,
    }

    rec = apply_customer_fixes(rec, full_text)
    return rec, None


# ══════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════

OUTPUT_COLUMNS = [
    "S No", "Customer Name", "Location", "Model", "Quantity",
    "PO Number", "PO Date", "Contact Person", "Contact Number",
    "Email ID", "Source File"
]

HEADER_COLOR  = "1F4E79"
ALT_ROW_COLOR = "D6E4F0"

def style_sheet(ws):
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = border
            if row_idx == 1:
                cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
                cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[1].height = 30
            else:
                cell.font      = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", start_color=ALT_ROW_COLOR)

    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"


def build_excel(pdf_files, output_path="extracted_data.xlsx"):
    all_records = []
    failed = []
    skipped = []

    for pdf_path in pdf_files:
        if not os.path.isfile(pdf_path):
            print(f"  [SKIP] Not found: {Path(pdf_path).name}")
            continue
        print(f"  Processing: {Path(pdf_path).name}")
        try:
            rec, err = extract_po_data(pdf_path)
        except Exception as e:
            err = f"Unexpected error: {e}"
            rec = None

        if err:
            if err.startswith("SKIPPED"):
                print(f"  [SKIPPED] {err}")
                skipped.append(f"{Path(pdf_path).name} — {err}")
            elif err.startswith("UNEXTRACTED"):
                print(f"  [UNEXTRACTED] {err}")
                failed.append(f"{Path(pdf_path).name} — no readable text")
            else:
                print(f"  [ERROR] {err}")
                failed.append(f"{Path(pdf_path).name} — {err}")
            continue
        if rec:
            all_records.append(rec)
            print(f"    ✓ Customer: {rec.get('Customer Name') or '—'}  |  PO: {rec.get('PO Number') or '—'}  |  Model: {rec.get('Model') or '—'}  |  Qty: {rec.get('Quantity') or '—'}")

    if not all_records:
        print("No data extracted.")
        return

    by_year = defaultdict(list)
    for rec in all_records:
        by_year[rec.get("PO Year", "Unclassified") or "Unclassified"].append(rec)

    sorted_years = sorted(
        (y for y in by_year if y != "Unclassified"), key=lambda y: int(y))
    if "Unclassified" in by_year:
        sorted_years.append("Unclassified")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for year in sorted_years:
            rows = by_year[year]
            for i, r in enumerate(rows, start=1):
                r["S No"] = i
            df = pd.DataFrame(rows)
            cols = [c for c in OUTPUT_COLUMNS if c in df.columns]
            df[cols].to_excel(writer, sheet_name=str(year)[:31], index=False)
            print(f"  Sheet '{year}': {len(df)} row(s)")

    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(output_path)

    print(f"\n✅ Done! Saved to: {output_path}")
    print(f"   Total processed: {len(all_records)} PO(s)")

    if skipped:
        print(f"\n⏭️  {len(skipped)} file(s) intentionally skipped (non-PO documents):")
        for f in skipped:
            print(f"    - {f}")

    if failed:
        print(f"\n⚠️  {len(failed)} file(s) could not be processed:")
        for f in failed:
            print(f"    - {f}")


# ══════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════

def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract PO PDFs → Excel (year-wise sheets, 10 key columns).")
    parser.add_argument("pdfs", nargs="*", help="PDF file paths")
    parser.add_argument("--folder", "-f", help="Folder containing PDF files")
    parser.add_argument("--output", "-o", default="extracted_data.xlsx")
    return parser.parse_args()

if __name__ == "__main__":
    args = parse_args()
    pdf_files = list(args.pdfs)
    if args.folder:
        pdf_files += sorted(str(p) for p in Path(args.folder).glob("*.pdf"))
        pdf_files += sorted(str(p) for p in Path(args.folder).glob("*.PDF"))
    if not pdf_files:
        print("Usage: python3 pdf_to_excel.py file1.pdf file2.pdf")
        print("       python3 pdf_to_excel.py --folder ./po_pdfs")
        sys.exit(1)
    print(f"\nFound {len(pdf_files)} PDF(s)...\n")
    build_excel(pdf_files, output_path=args.output)